"""封装获取excel数据"""
import openpyxl


class ExcelHandler:
    """读取excel数据"""
    def __init__(self, file_path):
        """初始化__init__"""
        self.file_path = file_path

    def read_data(self, sheet_name):
        """读取数据"""
        wb = openpyxl.open(self.file_path)
        ws = wb[sheet_name]
        cell_data = list(ws.values)
        data_lst = []
        keys = cell_data[0]
        # enumerate同时获取索引与值
        for sub_data in cell_data[1:]:
            data_dict = {}
            for ind, value in enumerate(sub_data):
                data_dict[keys[ind]] = value
            data_lst.append(data_dict)
        return data_lst
        # zip函数-元组与列表均可以
        # for value in cell_data[1:]:
        #     data_dict = dict(zip(keys, value))
        #     data_lst.append(data_dict)
        # return data_lst

    def write_data(self, sheet_name, row, column, data):
        workbook = openpyxl.load_workbook(self.file_path)
        worksheet = workbook[sheet_name]
        worksheet.cell(row=row, column=column).value = data
        workbook.save(self.file_path)
        workbook.close()


# if __name__ == "__main__":
#     p2p_excel_read = ExcelHandler(file_path=data_file_path)
#     excel_data = p2p_excel_read.read_data(sheet_name="register_test")
#     print(excel_data)
# 获取workbook
# workbook = openpyxl.open(data_file_path)
# print(workbook)

# 获取sheet
# work_sheet = workbook['register_test']
# print(work_sheet)

# 获取cell
# cell_data = work_sheet.cell(row=2, column=3)
# print(cell_data)

# 获取cell值
# print(cell_data.value)

# 获取某一行
# print(work_sheet[2])

# 获取所有行
# pprint(list(work_sheet.rows))

# 获取所有值
# pprint(list(work_sheet.values))


# # 打开文件
#
# workbook = openpyxl.load_workbook("list.xlsx")
#
# # 获取sheet
#
# worksheet = workbook["abc"]
#
# # 写入数据
#
# worksheet.cell(row=1, column=2).value = '哈哈哈哈'
#
# # 保存文件
# workbook.save("list.xlsx")
#
# # 关闭文件
# workbook.close()
